import os
import sqlite3
import sys
import time

from PyQt5 import QtCore
import pandas as pd
import xlsxwriter as xw
from PyQt5.QtCore import QSize, Qt, QDate
from PyQt5.QtGui import QPixmap, QIcon
from PyQt5.QtWidgets import QWidget, QApplication, QPushButton, QVBoxLayout, \
    QHBoxLayout, QMainWindow, QMessageBox, qApp, QAction, QTextEdit, QFileDialog, QInputDialog, QLabel, QProgressBar, \
    QLineEdit, QFrame, QCalendarWidget
from apyori import apriori
from openpyxl import load_workbook
from PyQt5 import QtGui
from xlrd import open_workbook
from collections import Counter


class Analyse(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.baglanti_olustur()

    def baglanti_olustur(self):
        self.baglanti = sqlite3.connect("Market-Basket.db")

        self.cursor = self.baglanti.cursor()

        sorgu = "Create Table If not exists Results(Produkts TEXT NOT NULL UNIQUE,Support REAL,Confidence REAL,Lift REAL)"

        self.cursor.execute(sorgu)
        self.baglanti.commit()

    def baglanti_kes(self):
        self.baglanti.close()

    def init_ui(self):
        self.veri_duzenle = QPushButton("Veri Düzenle")
        self.analiz = QPushButton("Analiz")
        self.bestseller = QPushButton("En Çok Satılanlar")
        self.calendar = QCalendarWidget(self)


        #####CSS AYARLAR####
        css = ("""
            color: #eff0f1;
            background-color: qlineargradient(x1: 0.5, y1: 0.5 x2: 0.5, y2: 1, stop: 0 #3b4045, stop: 0.5 #31363b);
            border-width: 1ex;
            border-color: black;
            border-style: solid;
            padding: 1ex;
            border-radius: 1ex;
            outline: none;
         
        """)
        css2 = ("""
            border-width: 0ex;
            border-color: black;
            border-style: solid;
            padding: 0ex;
            border-radius: 0ex;
            outline: none; 
            background: transparent;
        
        """)

        css3= ("""
            border-width: 0.5ex;
            border-color: black;
            border-style: solid;
            padding: 0.5ex;
            border-radius: 0.5ex;
            outline: none; 
        
        """)


        ###AYARLAR####
        self.veri_duzenle.setStyleSheet(css)
        self.analiz.setStyleSheet(css)
        self.bestseller.setStyleSheet(css)
        self.calendar.setStyleSheet(css3)
        self.calendar.setMaximumSize(300,300)

        self.setStyleSheet("background-color: white")
        labelImage = QLabel()
        pixmap = QPixmap("logo2.jpg")
        labelImage.setPixmap(pixmap)
        labelImage.setStyleSheet(css2)
        labelImage.setMaximumSize(400,400)
        v_box = QVBoxLayout()

        v_box.addWidget(self.calendar)

        v_box.addWidget(self.veri_duzenle)
        v_box.addWidget(self.analiz)
        v_box.addWidget(self.bestseller)
        v_box.addStretch(10)
        h_box = QHBoxLayout()
        h_box.addWidget(labelImage)

        h_box.addLayout(v_box)
        self.setLayout(h_box)


        self.setWindowTitle("Market-Basket Analyse")

        self.veri_duzenle.clicked.connect(self.excel_duzenleme)
        self.analiz.clicked.connect(self.analiz_yap)
        self.bestseller.clicked.connect(self.show_bestseller)

    def message(self):

        msg = QMessageBox()
        msg.setWindowTitle("Bilgilendirme")
        msg.setText("İşlem sürmektedir...")
        msg.setIcon(QMessageBox.Information)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setDefaultButton(QMessageBox.Ok)
        msg.exec_()


    def show_bestseller(self):
        dosya_ismi = QFileDialog.getOpenFileName(self, "Dosya Aç", os.getenv("HOME"))

        excel_file = pd.read_excel(dosya_ismi[0])

        my_list = excel_file['ProduktHierarchie'].tolist()

        sonuc = Counter(my_list)

        bestseller = sonuc.most_common(10)


        #DataBase#
        conn = sqlite3.connect("Market-Basket.db")
        cur = conn.cursor()

        cur.execute("Create Table If not exists Bestseller(ProduktHierarchie TEXT NOT NULL UNIQUE,Satılan_Adet REAL)")


        sql_statement = "INSERT INTO Bestseller VALUES(?,?)"

        cur.executemany(sql_statement, bestseller)
        conn.commit()

        # Popup#
        msg = QMessageBox()
        msg.setWindowTitle("Bilgilendirme")
        msg.setText("En Çok Satılanlar Listesi Veritabanında Oluşturuldu!")
        msg.setIcon(QMessageBox.Information)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setDefaultButton(QMessageBox.Ok)
        msg.exec_()


    def excel_duzenleme(self):
        # Kullanıcı tarafından belirtilen dosyanın açılması için gerekli kod yapısı
        dosya_ismi = QFileDialog.getOpenFileName(self, "Dosya Aç", os.getenv("HOME"))
        self.message()
        a = load_workbook(dosya_ismi[0])

        wb = open_workbook(dosya_ismi[0])

        sayfa1 = wb.sheet_by_index(0)

        produkt_hierarchie = []
        lieferung_nummer = []

        m = 0
        k = 0

        for satir_no in range(sayfa1.nrows):
            produkt_hierarchie.append(sayfa1.cell(satir_no, 4).value)
            lieferung_nummer.append(sayfa1.cell(satir_no, 2).value)

        a.close()

        workbook = xw.Workbook("Analyse.xlsx")
        daten_sheet = workbook.add_worksheet(name='Daten')

        for i in range(len(lieferung_nummer) - 1):
            if lieferung_nummer[i] == lieferung_nummer[i + 1]:
                daten_sheet.write(m, 0, lieferung_nummer[i])
                if produkt_hierarchie[i] != produkt_hierarchie[i + 1]:
                    daten_sheet.write(m, k + 1, produkt_hierarchie[i])
                    k = k + 1
                else:
                    daten_sheet.write(m, k + 1, produkt_hierarchie[i])
            else:
                daten_sheet.write(m, 0, lieferung_nummer[i])
                daten_sheet.write(m, k + 1, produkt_hierarchie[i])
                k = 0
                m = m + 1

        workbook.close()

        #Popup#
        msg = QMessageBox()
        msg.setWindowTitle("Bilgilendirme")
        msg.setText("Excel Dosyası Başarıyla güncellendi!")
        msg.setIcon(QMessageBox.Information)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setDefaultButton(QMessageBox.Ok)
        msg.exec_()


    def analiz_yap(self):
        dosya_ismi = QFileDialog.getOpenFileName(self, "Dosya Aç", os.getenv("HOME"))
        self.message()
        excel_file = pd.ExcelFile(dosya_ismi[0])
        market_basket = excel_file.parse('Daten', header=None)

        df = market_basket.drop([0], axis=0)

        daten = df.drop([0], axis=1)


        total_records = len(daten)

        # Apriori algoritması devreye sokulmaktadır....
        records = []

        for i in range(0, total_records):
            records.append([str(daten.values[i, j]) for j in range(0, daten.shape[1])])

        associationRules = apriori(records, min_support=0.0050 , min_lift=3, min_confidence=0.7 , min_length=2)

        rules = associationRules

        self.veri_aktar(rules)
        msg = QMessageBox()
        msg.setWindowTitle("Bilgilendirme")
        msg.setText("Birliktelik Kuralları Oluşturuldu!")
        msg.setIcon(QMessageBox.Information)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setDefaultButton(QMessageBox.Ok)
        msg.exec_()


    def deger_ekle(self, produkts, support, confidence, lift):
        sorgu = "INSERT INTO Results VALUES(?,?,?,?)"
        try:
            self.cursor.execute(sorgu, (produkts, support, confidence, lift))
            self.baglanti.commit()

        except:
            pass

    def veri_aktar(self,rules):

        for item in rules:
            pair = item[0]
            items = [x for x in pair]

            # Bu kısımda 4 farklı parametreye değerler atanmaktadır.
            produkts = str(items[0] + "->" + items[1])
            support = str(item[1])
            confidence = str(item[2][0][2])
            lift = str(item[2][0][3])
            if "nan" in produkts:
                pass

            else:
                self.deger_ekle(produkts, support, confidence, lift)

        msg = QMessageBox()
        msg.setWindowTitle("Bilgilendirme")
        msg.setText("Birliktelik Kuralları Veritabanına Eklendi!")
        msg.setIcon(QMessageBox.Information)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setDefaultButton(QMessageBox.Ok)
        msg.exec_()

class Menu(QMainWindow):

    def __init__(self):

        super().__init__()

        self.pencere = Analyse()
        self.pencere.setGeometry(500, 500, 200, 200)
        self.setCentralWidget(self.pencere)

        self.menuleri_olustur()
        self.init_ui()

    def init_ui(self):

        self.setGeometry(500, 500,500, 500)
        css = """
            color: black;
            background-color: grey;
            selection-background-color:#3daee9;
            selection-color: yellow;
            background-clip: border;
            border-image: none;
            border: 0px transparent black;
            outline: 0;
        
        """
        self.setStyleSheet(css)
        self.setWindowIcon(QIcon("icon.png"))
        self.setMinimumSize(QSize(650,450))
        self.setMaximumSize(QSize(800,800))
        self.show()

    def menuleri_olustur(self):

        menubar = self.menuBar()

        file = menubar.addMenu("File")

        veri_duzenle = QAction("Veri Düzenle", self)
        veri_duzenle.setShortcut("Ctrl+O")

        analiz = QAction("Analiz",self)
        analiz.setShortcut("Ctrl+A")

        cikis = QAction("Çıkış", self)

        cikis.setShortcut("Ctrl+Q")

        file.addAction(veri_duzenle)
        file.addAction(analiz)
        file.addAction(cikis)

        file.triggered.connect(self.response)

        self.setWindowTitle("MARKET-BASKET ANALYSE")

        self.show()

    def response(self, action):

        if action.text() == "Veri Düzenle":
            self.pencere.dosya_ac()

        elif action.text() == "Analiz":
            self.pencere.analiz_yap()

        elif action.text() == "Çıkış":
            qApp.quit()

stylesheet = """
    QLabel {
    border: 2px solid green;
    border-radius: 4px;
    padding: 2px;
    font-size: 15px;
    font-family: Siemens Sans SC;
}
"""


app = QApplication(sys.argv)
app.setStyleSheet(stylesheet)
menu = Menu()

sys.exit(app.exec_())
